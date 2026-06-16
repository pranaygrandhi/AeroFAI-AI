"""
AS9102 Form Generation Module

This module generates aerospace First Article Inspection (FAI) documentation.
- Form 1: First Article Inspection Report (FAIR)
- Form 2: Inspection Planning (planned approach)
- Form 3: Characteristics Accountability (inspection results)
"""

from typing import List, Optional, Dict, Any
from datetime import datetime
import json
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

import logging

logger = logging.getLogger(__name__)


class AS9102Generator:
    """Generate AS9102 forms in Excel format."""

    def __init__(self):
        self.workbook = None
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_form_3(
        self,
        drawing_id: int,
        part_name: str,
        part_number: str,
        revision: str,
        serial_number: str,
        characteristics: List[Dict[str, Any]],
        inspection_date: Optional[datetime] = None,
        inspector_name: Optional[str] = None,
        customer: Optional[str] = None,
        supplier: Optional[str] = None,
    ) -> BytesIO:
        """
        Generate AS9102 Form 3: Characteristics Accountability (Inspection Results).

        Args:
            drawing_id: Drawing ID
            part_name: Part name/description
            part_number: Part number
            revision: Drawing revision
            serial_number: Part serial number
            characteristics: List of characteristic dicts with measured values and status
            inspection_date: Date of inspection
            inspector_name: Name of inspector
            customer: Customer name
            supplier: Supplier name

        Returns:
            BytesIO object containing Excel file
        """
        if Workbook is None:
            logger.error("openpyxl not installed. Cannot generate forms.")
            raise ImportError("openpyxl is required for Form generation")

        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Form 3 - Characteristics"

        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 20

        # Header section
        current_row = 1
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        ws["A1"] = "AS9102 FORM 3 - FIRST ARTICLE INSPECTION CHARACTERISTICS"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        # Part identification section
        current_row = 3
        self._add_section_header(ws, current_row, "PART IDENTIFICATION", header_fill, header_font)
        current_row += 1

        # Part info grid
        self._add_info_row(ws, current_row, "Part Name:", part_name)
        current_row += 1
        self._add_info_row(ws, current_row, "Part Number:", part_number)
        current_row += 1
        self._add_info_row(ws, current_row, "Revision:", revision)
        current_row += 1
        self._add_info_row(ws, current_row, "Serial Number:", serial_number)
        current_row += 1
        self._add_info_row(ws, current_row, "Customer:", customer or "")
        current_row += 1
        self._add_info_row(ws, current_row, "Supplier:", supplier or "")
        current_row += 1
        self._add_info_row(ws, current_row, "Inspection Date:", inspection_date.strftime("%Y-%m-%d") if inspection_date else "")
        current_row += 1
        self._add_info_row(ws, current_row, "Inspector:", inspector_name or "")
        current_row += 2

        # Characteristics table header
        self._add_section_header(ws, current_row, "INSPECTION CHARACTERISTICS", header_fill, header_font)
        current_row += 1

        # Table headers
        headers = [
            "Balloon #",
            "Characteristic/GD&T",
            "Type",
            "Nominal",
            "Upper Limit",
            "Lower Limit",
            "Unit",
            "Measured Value",
            "Status",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.thin_border

        current_row += 1

        # Add characteristics data
        pass_count = 0
        fail_count = 0

        for char in characteristics:
            cells = [
                ws.cell(row=current_row, column=1),
                ws.cell(row=current_row, column=2),
                ws.cell(row=current_row, column=3),
                ws.cell(row=current_row, column=4),
                ws.cell(row=current_row, column=5),
                ws.cell(row=current_row, column=6),
                ws.cell(row=current_row, column=7),
                ws.cell(row=current_row, column=8),
                ws.cell(row=current_row, column=9),
            ]

            values = [
                char.get("balloon_no", ""),
                char.get("requirement", ""),
                char.get("type", ""),
                char.get("nominal", ""),
                char.get("upper_limit", ""),
                char.get("lower_limit", ""),
                char.get("unit", ""),
                char.get("measured_value", ""),
                char.get("status", "pending"),
            ]

            for cell, value in zip(cells, values):
                cell.value = value
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Color code status
                status = char.get("status", "pending").lower()
                if status == "pass":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    pass_count += 1
                elif status == "fail":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    fail_count += 1

            current_row += 1

        # Summary section
        current_row += 1
        self._add_section_header(ws, current_row, "INSPECTION SUMMARY", header_fill, header_font)
        current_row += 1

        total_chars = len(characteristics)
        self._add_info_row(ws, current_row, "Total Characteristics:", total_chars)
        current_row += 1
        self._add_info_row(ws, current_row, "Passed:", pass_count)
        current_row += 1
        self._add_info_row(ws, current_row, "Failed:", fail_count)
        current_row += 1
        self._add_info_row(ws, current_row, "Pending:", total_chars - pass_count - fail_count)
        current_row += 1

        result_status = "PASS" if fail_count == 0 else "FAIL"
        result_color = "C6EFCE" if fail_count == 0 else "FFC7CE"

        ws[f"A{current_row}"] = "RESULT:"
        ws[f"B{current_row}"] = result_status
        ws[f"B{current_row}"].fill = PatternFill(start_color=result_color, end_color=result_color, fill_type="solid")
        ws[f"B{current_row}"].font = Font(bold=True, size=12)

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def generate_form_1(
        self,
        drawing_id: int,
        part_name: str,
        part_number: str,
        revision: str,
        supplier: str,
        customer: str,
        inspection_level: str = "Comprehensive",
        special_requirements: Optional[str] = None,
    ) -> BytesIO:
        """
        Generate AS9102 Form 1: First Article Inspection Report (FAIR).

        Args:
            drawing_id: Drawing ID
            part_name: Part name
            part_number: Part number
            revision: Drawing revision
            supplier: Supplier name
            customer: Customer name
            inspection_level: Level of inspection
            special_requirements: Any special requirements

        Returns:
            BytesIO object containing Excel file
        """
        if Workbook is None:
            raise ImportError("openpyxl is required for Form generation")

        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Form 1 - FAIR"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Header
        ws["A1"] = "AS9102 FORM 1 - FIRST ARTICLE INSPECTION REPORT (FAIR)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        # Part Information
        current_row = 3
        self._add_section_header(ws, current_row, "PART IDENTIFICATION", header_fill, header_font)
        current_row += 1

        self._add_form_field(ws, current_row, "Part Name:", part_name)
        current_row += 1
        self._add_form_field(ws, current_row, "Part Number:", part_number)
        current_row += 1
        self._add_form_field(ws, current_row, "Revision:", revision)
        current_row += 1
        self._add_form_field(ws, current_row, "Supplier:", supplier)
        current_row += 1
        self._add_form_field(ws, current_row, "Customer:", customer)
        current_row += 2

        # Inspection Details
        self._add_section_header(ws, current_row, "INSPECTION PLAN", header_fill, header_font)
        current_row += 1

        self._add_form_field(ws, current_row, "Inspection Level:", inspection_level)
        current_row += 1
        self._add_form_field(ws, current_row, "Special Requirements:", special_requirements or "None")
        current_row += 1
        self._add_form_field(ws, current_row, "Inspection Date:", datetime.now().strftime("%Y-%m-%d"))
        current_row += 2

        # Approval section
        self._add_section_header(ws, current_row, "APPROVALS", header_fill, header_font)
        current_row += 1
        self._add_form_field(ws, current_row, "Prepared By:", "_" * 30)
        current_row += 1
        self._add_form_field(ws, current_row, "Approved By:", "_" * 30)
        current_row += 1
        self._add_form_field(ws, current_row, "Date:", "_" * 30)

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def generate_form_2(
        self,
        drawing_id: int,
        part_name: str,
        part_number: str,
        revision: str,
        characteristics: List[Dict[str, Any]],
        supplier: str,
        customer: str,
    ) -> BytesIO:
        """
        Generate AS9102 Form 2: Inspection Planning.

        Args:
            drawing_id: Drawing ID
            part_name: Part name
            part_number: Part number
            revision: Drawing revision
            characteristics: List of characteristics to be inspected
            supplier: Supplier name
            customer: Customer name

        Returns:
            BytesIO object containing Excel file
        """
        if Workbook is None:
            raise ImportError("openpyxl is required for Form generation")

        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Form 2 - Planning"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Header
        ws["A1"] = "AS9102 FORM 2 - INSPECTION PLANNING"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:H1")

        # Part Information
        current_row = 3
        self._add_section_header(ws, current_row, "PART IDENTIFICATION", header_fill, header_font)
        current_row += 1

        self._add_form_field(ws, current_row, "Part Name:", part_name)
        current_row += 1
        self._add_form_field(ws, current_row, "Part Number:", part_number)
        current_row += 1
        self._add_form_field(ws, current_row, "Revision:", revision)
        current_row += 1
        self._add_form_field(ws, current_row, "Supplier:", supplier)
        current_row += 1
        self._add_form_field(ws, current_row, "Customer:", customer)
        current_row += 2

        # Inspection plan table
        self._add_section_header(ws, current_row, "INSPECTION CHARACTERISTICS PLAN", header_fill, header_font)
        current_row += 1

        headers = ["Balloon #", "Characteristic", "Type", "Method", "Tool", "Frequency", "Acceptance Criteria"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = self.thin_border

        current_row += 1

        for char in characteristics:
            for col_num, value in enumerate(
                [
                    char.get("balloon_no", ""),
                    char.get("requirement", ""),
                    char.get("type", ""),
                    "Visual" if char.get("type") == "visual" else "Instrument",
                    char.get("tool", "CMM"),
                    "100%",
                    f"Nominal: {char.get('nominal', '')}",
                ],
                1,
            ):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = self.thin_border

            current_row += 1

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output

    def _add_section_header(self, ws, row, title, fill, font):
        """Add a section header."""
        ws[f"A{row}"] = title
        ws[f"A{row}"].fill = fill
        ws[f"A{row}"].font = font
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

    def _add_info_row(self, ws, row, label, value):
        """Add an info row with label and value."""
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        ws.merge_cells(f"B{row}:C{row}")

    def _add_form_field(self, ws, row, label, value):
        """Add a form field."""
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].border = Border(bottom=Side(style="thin"))
        ws.merge_cells(f"B{row}:F{row}")
